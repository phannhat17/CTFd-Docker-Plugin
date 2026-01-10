"""
Admin Routes - Container management for admins
"""
from flask import Blueprint, request, jsonify, render_template, session
from CTFd.utils.decorators import admins_only
from CTFd.models import db
from ..models.instance import ContainerInstance
from ..models.challenge import ContainerChallenge
from ..models.flag import ContainerFlagAttempt
from ..models.audit import ContainerAuditLog
from ..models.config import ContainerConfig
from CTFd.utils.security.auth import generate_nonce

admin_bp = Blueprint('containers_admin', __name__, url_prefix='/admin/containers')

# Template filters
@admin_bp.app_template_filter('get_user')
def get_user_filter(user_id):
    """Get user by ID"""
    from CTFd.models import Users
    return Users.query.filter_by(id=user_id).first()

@admin_bp.app_template_filter('get_team')
def get_team_filter(team_id):
    """Get team by ID"""
    from CTFd.models import Teams
    return Teams.query.filter_by(id=team_id).first()

# Global services
docker_service = None
container_service = None
anticheat_service = None


def set_services(d_service, c_service, a_service):
    """Inject services"""
    global docker_service, container_service, anticheat_service
    docker_service = d_service
    container_service = c_service
    anticheat_service = a_service


# ============================================================================
# Admin Pages
# ============================================================================

def _get_docker_status():
    """Helper function to get Docker status for all pages"""
    connected = False
    docker_info = None
    
    try:
        if docker_service and docker_service.is_connected():
            connected = True
            client = docker_service.client
            version_info = client.version()
            system_info = client.info()
            
            docker_info = {
                'version': version_info.get('Version', 'Unknown'),
                'api_version': version_info.get('ApiVersion', 'Unknown'),
                'containers_running': system_info.get('ContainersRunning', 0),
                'containers_stopped': system_info.get('ContainersStopped', 0),
                'images': system_info.get('Images', 0),
                'cpus': system_info.get('NCPU', 0),
                'memory_total': system_info.get('MemTotal', 0)
            }
    except:
        pass
    
    return connected, docker_info


@admin_bp.route('/dashboard')
@admins_only
def dashboard():
    """Admin dashboard - overview of all containers"""
    from CTFd.utils import get_config
    from CTFd.models import Users, Teams
    
    # Fetch instances grouped by status
    running_instances = ContainerInstance.query.filter_by(status='running').order_by(ContainerInstance.created_at.desc()).all()
    provisioning_instances = ContainerInstance.query.filter_by(status='provisioning').order_by(ContainerInstance.created_at.desc()).all()
    solved_instances = ContainerInstance.query.filter_by(status='solved').order_by(ContainerInstance.created_at.desc()).limit(20).all()
    stopped_instances = ContainerInstance.query.filter_by(status='stopped').order_by(ContainerInstance.created_at.desc()).limit(20).all()
    error_instances = ContainerInstance.query.filter_by(status='error').order_by(ContainerInstance.created_at.desc()).limit(10).all()
    
    # Get stats
    running_count = len(running_instances)
    total_count = ContainerInstance.query.count()
    
    # Get Docker status
    connected, docker_info = _get_docker_status()
    
    # Check if teams mode
    is_teams_mode = get_config('user_mode') == 'teams'
    
    return render_template('container_dashboard.html',
                         running_instances=running_instances,
                         provisioning_instances=provisioning_instances,
                         solved_instances=solved_instances,
                         stopped_instances=stopped_instances,
                         error_instances=error_instances,
                         running_count=running_count,
                         total_count=total_count,
                         connected=connected,
                         docker_info=docker_info,
                         is_teams_mode=is_teams_mode,
                         active_page='dashboard')


@admin_bp.route('/settings')
@admins_only
def settings():
    """Settings page"""
    # Get all config values
    settings_data = {
        'docker_type': ContainerConfig.get('docker_type', 'local'),
        'ssh_hostname': ContainerConfig.get('ssh_hostname', ''),
        'ssh_port': ContainerConfig.get('ssh_port', '22'),
        'ssh_user': ContainerConfig.get('ssh_user', 'root'),
        'ssh_key_content': ContainerConfig.get('ssh_key_content', ''),
        'ssh_known_hosts': ContainerConfig.get('ssh_known_hosts', ''),
        'docker_base_url': ContainerConfig.get('docker_socket', ''),
        'docker_hostname': ContainerConfig.get('connection_host', ''),
        'container_expiration': ContainerConfig.get('default_timeout', '60'),
        'max_renewals': ContainerConfig.get('max_renewals', '3'),
        'container_maxmemory': ContainerConfig.get('max_memory', '512m'),
        'container_maxcpu': ContainerConfig.get('max_cpu', '0.5'),
        'port_range_start': ContainerConfig.get('port_range_start', '30000'),
        'port_range_end': ContainerConfig.get('port_range_end', '31000'),
        # Subdomain routing (Traefik)
        'subdomain_enabled': ContainerConfig.get('subdomain_enabled', 'false'),
        'subdomain_base_domain': ContainerConfig.get('subdomain_base_domain', ''),
        'subdomain_network': ContainerConfig.get('subdomain_network', 'ctfd-network'),
    }
    
    # Get Docker status
    connected, docker_info = _get_docker_status()
    error_message = None
    
    return render_template('container_settings.html', 
                         settings=settings_data,
                         connected=connected,
                         docker_info=docker_info,
                         error_message=error_message,
                         active_page='settings')


@admin_bp.route('/cheats')
@admins_only
def cheats():
    """Cheat detection logs"""
    from CTFd.models import Users, Teams
    
    # Get all cheat attempts (flag reuse) with relationships loaded
    cheat_logs = ContainerFlagAttempt.query.filter(
        ContainerFlagAttempt.is_cheating == True
    ).order_by(ContainerFlagAttempt.timestamp.desc()).all()
    
    # Load owner info (user or team) for each log
    for log in cheat_logs:
        # Get submitter info
        submitter_user = Users.query.filter_by(id=log.user_id).first()
        if submitter_user:
            log.submitter_team = Teams.query.filter_by(id=submitter_user.team_id).first() if submitter_user.team_id else None
            log.submitter_user_obj = submitter_user
        
        # Get flag owner info
        if log.flag_owner_account_id:
            # Try to find owner (could be user or team)
            owner_user = Users.query.filter_by(id=log.flag_owner_account_id).first()
            if owner_user:
                log.owner_team = Teams.query.filter_by(id=owner_user.team_id).first() if owner_user.team_id else None
                log.owner_user_obj = owner_user
            else:
                # Might be team account
                owner_team = Teams.query.filter_by(id=log.flag_owner_account_id).first()
                if owner_team:
                    log.owner_team = owner_team
                    log.owner_user_obj = None
    
    # Get Docker status
    connected, docker_info = _get_docker_status()
    
    return render_template('container_cheat.html', 
                         cheat_logs=cheat_logs, 
                         connected=connected, 
                         docker_info=docker_info,
                         active_page='cheats')


# ============================================================================
# Admin APIs
# ============================================================================

@admin_bp.route('/api/instances', methods=['GET'], endpoint='api_instances')
@admins_only
def api_instances():
    """
    List all container instances
    
    Query params:
        status: Filter by status
        challenge_id: Filter by challenge
        account_id: Filter by account
        limit: Limit results (default 100)
    """
    try:
        query = ContainerInstance.query
        
        # Filters
        if request.args.get('status'):
            query = query.filter_by(status=request.args.get('status'))
        if request.args.get('challenge_id'):
            query = query.filter_by(challenge_id=request.args.get('challenge_id'))
        if request.args.get('account_id'):
            query = query.filter_by(account_id=request.args.get('account_id'))
        
        limit = min(int(request.args.get('limit', 100)), 500)
        
        instances = query.order_by(ContainerInstance.created_at.desc()).limit(limit).all()
        
        # Serialize
        result = []
        for instance in instances:
            challenge = ContainerChallenge.query.get(instance.challenge_id)
            result.append({
                'id': instance.id,
                'uuid': instance.uuid,
                'challenge_id': instance.challenge_id,
                'challenge_name': challenge.name if challenge else 'Unknown',
                'account_id': instance.account_id,
                'container_id': instance.container_id,
                'port': instance.connection_port,
                'status': instance.status,
                'created_at': instance.created_at.isoformat(),
                'expires_at': instance.expires_at.isoformat(),
                'stopped_at': instance.stopped_at.isoformat() if instance.stopped_at else None,
                'renewal_count': instance.renewal_count
            })
        
        return jsonify({'instances': result})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/instances/<int:instance_id>', methods=['DELETE'], endpoint='api_delete_instance')
@admins_only
def delete_instance(instance_id):
    """Delete a specific instance"""
    try:
        instance = ContainerInstance.query.get(instance_id)
        if not instance:
            return jsonify({'error': 'Instance not found'}), 404
        
        # Stop container if running
        if instance.status == 'running' and instance.container_id:
            container_service.stop_instance(instance, user_id=None, reason='admin_delete')
        
        # Delete from database
        db.session.delete(instance)
        db.session.commit()
        
        return jsonify({'success': True})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/instances/<int:instance_id>/stop', methods=['POST'], endpoint='api_stop_instance')
@admins_only
def stop_instance(instance_id):
    """Stop a specific instance"""
    try:
        instance = ContainerInstance.query.get(instance_id)
        if not instance:
            return jsonify({'error': 'Instance not found'}), 404
        
        success = container_service.stop_instance(instance, user_id=None, reason='admin')
        
        if success:
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Failed to stop instance'}), 500
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/instances/<int:instance_id>/logs', methods=['GET'], endpoint='api_instance_logs')
@admins_only
def get_instance_logs(instance_id):
    """Get container logs"""
    try:
        instance = ContainerInstance.query.get(instance_id)
        if not instance:
            return jsonify({'error': 'Instance not found'}), 404
        
        if not instance.container_id:
            return jsonify({'error': 'No container ID'}), 404
        
        logs = docker_service.get_container_logs(instance.container_id, tail=500)
        
        return jsonify({'logs': logs})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/bulk-delete', methods=['POST'], endpoint='api_bulk_delete')
@admins_only
def api_bulk_delete():
    """Bulk delete instances"""
    try:
        data = request.get_json()
        instance_ids = data.get('instance_ids', [])
        
        if not instance_ids:
            return jsonify({'error': 'No instance IDs provided'}), 400
        
        deleted_count = 0
        for instance_id in instance_ids:
            instance = ContainerInstance.query.get(instance_id)
            if instance:
                # Stop container if running
                if instance.status == 'running' and instance.container_id:
                    container_service.stop_instance(instance, user_id=None, reason='admin_bulk_delete')
                
                db.session.delete(instance)
                deleted_count += 1
        
        db.session.commit()
        
        return jsonify({'success': True, 'deleted': deleted_count})
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/stats', methods=['GET'], endpoint='api_stats')
@admins_only
def get_stats():
    """Get statistics"""
    try:
        stats = {
            'total_instances': ContainerInstance.query.count(),
            'running': ContainerInstance.query.filter_by(status='running').count(),
            'stopped': ContainerInstance.query.filter_by(status='stopped').count(),
            'solved': ContainerInstance.query.filter_by(status='solved').count(),
            'error': ContainerInstance.query.filter_by(status='error').count(),
            'total_attempts': ContainerFlagAttempt.query.count(),
            'cheat_attempts': ContainerFlagAttempt.query.filter_by(is_cheating=True).count(),
            'docker_connected': docker_service.is_connected() if docker_service else False
        }
        
        return jsonify(stats)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/cheats', methods=['GET'], endpoint='api_cheats')
@admins_only
def list_cheats():
    """List cheat attempts"""
    try:
        limit = min(int(request.args.get('limit', 100)), 500)
        
        attempts = anticheat_service.get_cheat_attempts(limit=limit)
        
        result = []
        for attempt in attempts:
            result.append({
                'id': attempt.id,
                'challenge_id': attempt.challenge_id,
                'account_id': attempt.account_id,
                'user_id': attempt.user_id,
                'flag_owner_account_id': attempt.flag_owner_account_id,
                'timestamp': attempt.timestamp.isoformat(),
                'ip_address': attempt.ip_address
            })
        
        return jsonify({'cheats': result})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/config', methods=['GET'], endpoint='api_config')
@admins_only
def get_config():
    """Get plugin configuration"""
    try:
        config = ContainerConfig.get_all()
        return jsonify({'config': config})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/config', methods=['POST'], endpoint='api_config_update')
@admins_only
def update_config():
    """
    Update plugin configuration and create SSH config if needed
    """
    try:
        import os
        data = request.get_json()
        
        # Update all config
        for key, value in data.items():
            ContainerConfig.set(key, str(value))
        
        # Handle Docker connection based on type
        docker_type = data.get('docker_type', 'local')
        
        if docker_type == 'ssh':
            # Create SSH config files
            ssh_dir = os.path.expanduser('~/.ssh')
            os.makedirs(ssh_dir, exist_ok=True)
            
            ssh_hostname = data.get('ssh_hostname', '')
            ssh_port = data.get('ssh_port', '22')
            ssh_user = data.get('ssh_user', 'root')
            ssh_key_content = data.get('ssh_key_content', '')
            ssh_known_hosts = data.get('ssh_known_hosts', '')
            
            if not ssh_hostname:
                return jsonify({'error': 'SSH hostname is required'}), 400
            
            # Create unique host alias
            host_alias = f"ctfd-docker-{ssh_hostname.replace('.', '-')}"
            
            # Write SSH private key
            key_path = os.path.join(ssh_dir, f'{host_alias}_key')
            if ssh_key_content:
                with open(key_path, 'w') as f:
                    f.write(ssh_key_content)
                os.chmod(key_path, 0o600)
            
            # Write/update known_hosts
            known_hosts_path = os.path.join(ssh_dir, 'known_hosts')
            if ssh_known_hosts:
                # Read existing known_hosts
                existing_hosts = []
                if os.path.exists(known_hosts_path):
                    with open(known_hosts_path, 'r') as f:
                        existing_hosts = f.readlines()
                
                # Remove old entries for this host
                existing_hosts = [line for line in existing_hosts if ssh_hostname not in line]
                
                # Add new entry
                existing_hosts.append(ssh_known_hosts.strip() + '\n')
                
                # Write back
                with open(known_hosts_path, 'w') as f:
                    f.writelines(existing_hosts)
                os.chmod(known_hosts_path, 0o644)
            
            # Create/update SSH config
            config_path = os.path.join(ssh_dir, 'config')
            
            # Read existing config
            existing_config = []
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    existing_config = f.readlines()
            
            # Remove old config for this host alias
            new_config = []
            skip_until_next_host = False
            for line in existing_config:
                if line.strip().startswith('Host ') and host_alias in line:
                    skip_until_next_host = True
                elif line.strip().startswith('Host ') and host_alias not in line:
                    skip_until_next_host = False
                
                if not skip_until_next_host:
                    new_config.append(line)
            
            # Add new config
            new_config.append('\n')
            new_config.append('# CTFd Docker Plugin - Auto-generated\n')
            new_config.append(f'Host {host_alias}\n')
            new_config.append(f'    HostName {ssh_hostname}\n')
            new_config.append(f'    User {ssh_user}\n')
            new_config.append(f'    Port {ssh_port}\n')
            if ssh_key_content:
                new_config.append(f'    IdentityFile {key_path}\n')
            new_config.append(f'    StrictHostKeyChecking yes\n')
            new_config.append(f'    UserKnownHostsFile {known_hosts_path}\n')
            
            # Write config
            with open(config_path, 'w') as f:
                f.writelines(new_config)
            os.chmod(config_path, 0o644)
            
            # Update docker socket URL to use SSH alias
            docker_socket = f'ssh://{host_alias}'
            ContainerConfig.set('docker_socket', docker_socket)
            
            # Reconnect Docker
            docker_service.base_url = docker_socket
            docker_service._connect()
        else:
            # Local docker
            docker_socket = 'unix://var/run/docker.sock'
            ContainerConfig.set('docker_socket', docker_socket)
            
            # Reconnect Docker
            docker_service.base_url = docker_socket
            docker_service._connect()
        
        return jsonify({'success': True})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/cleanup/expired', methods=['POST'], endpoint='api_cleanup_expired')
@admins_only
def cleanup_expired():
    """Manually trigger cleanup of expired instances"""
    try:
        container_service.cleanup_expired_instances()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/cleanup/old', methods=['POST'], endpoint='api_cleanup_old')
@admins_only
def cleanup_old():
    """Manually trigger cleanup of old instances"""
    try:
        container_service.cleanup_old_instances()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/images', methods=['GET'], endpoint='api_images')
@admins_only
def list_images():
    """List available Docker images"""
    try:
        if not docker_service:
            return jsonify({'error': 'Docker service not available'}), 500
        
        images = docker_service.list_images()
        # Extract image names/tags
        image_list = []
        for img in images:
            if img.tags:
                image_list.extend(img.tags)
        
        return jsonify({'images': image_list})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@admin_bp.route('/api/docker/health', methods=['GET'], endpoint='api_docker_health')
@admins_only
def docker_health_check():
    """
    Check Docker connection health
    
    Returns:
        {
            "connected": bool,
            "docker_version": str,
            "api_version": str,
            "server_info": {
                "containers": int,
                "images": int,
                "memory_total": int,
                "cpus": int
            },
            "error": str (if connection failed)
        }
    """
    try:
        if not docker_service:
            return jsonify({
                'connected': False,
                'error': 'Docker service not initialized'
            }), 500
        
        # Check connection
        is_connected = docker_service.is_connected()
        
        if not is_connected:
            return jsonify({
                'connected': False,
                'error': 'Cannot connect to Docker daemon',
                'socket': ContainerConfig.get('docker_socket', 'Not configured')
            })
        
        # Get Docker info
        try:
            client = docker_service.client
            version_info = client.version()
            system_info = client.info()
            
            return jsonify({
                'connected': True,
                'docker_version': version_info.get('Version', 'Unknown'),
                'api_version': version_info.get('ApiVersion', 'Unknown'),
                'server_info': {
                    'containers': system_info.get('Containers', 0),
                    'containers_running': system_info.get('ContainersRunning', 0),
                    'containers_paused': system_info.get('ContainersPaused', 0),
                    'containers_stopped': system_info.get('ContainersStopped', 0),
                    'images': system_info.get('Images', 0),
                    'memory_total': system_info.get('MemTotal', 0),
                    'cpus': system_info.get('NCPU', 0),
                    'server_version': system_info.get('ServerVersion', 'Unknown'),
                    'operating_system': system_info.get('OperatingSystem', 'Unknown'),
                    'architecture': system_info.get('Architecture', 'Unknown')
                },
                'socket': ContainerConfig.get('docker_socket', 'Not configured')
            })
        except Exception as info_error:
            return jsonify({
                'connected': True,
                'error': f'Connected but failed to get info: {str(info_error)}',
                'socket': ContainerConfig.get('docker_socket', 'Not configured')
            })
    
    except Exception as e:
        return jsonify({
            'connected': False,
            'error': str(e)
        }), 500


# ============================================================================
# Import Challenges from Excel
# ============================================================================

@admin_bp.route('/import', methods=['GET'])
@admins_only
def import_challenges_page():
    """Show import page"""
    # Get Docker status for template
    connected, docker_info = _get_docker_status()
    return render_template('container_import.html',
                         docker_connected=connected,
                         docker_info=docker_info)


@admin_bp.route('/api/import', methods=['POST'])
@admins_only
def import_challenges():
    """Import challenges from Excel file (API endpoint)"""
    # Get Docker status for error responses
    connected, docker_info = _get_docker_status()
    try:
        import openpyxl
        import re
        from werkzeug.utils import secure_filename
        from CTFd.models import Challenges, Flags
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': 'File must be Excel (.xlsx or .xls)'}), 400
        
        # Read Excel
        wb = openpyxl.load_workbook(file, data_only=True)
        
        # Try to find "Challenges" sheet or use first sheet
        if 'Challenges' in wb.sheetnames:
            ws = wb['Challenges']
        else:
            ws = wb.active
        
        # Parse header row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value.lower().strip() if cell.value else '')
        
        # Required columns
        required = ['name', 'category', 'image']
        for req in required:
            if req not in headers:
                return jsonify({'success': False, 'error': f'Missing required column: {req}'}), 400
        
        # Import challenges
        created_count = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Map row to dict
                row_data = {}
                for idx, value in enumerate(row):
                    if idx < len(headers) and headers[idx]:
                        row_data[headers[idx]] = value
                
                # Skip empty rows
                if not row_data.get('name'):
                    continue
                
                # Parse flag pattern
                flag_pattern = row_data.get('flag_pattern', 'CTF{flag}')
                flag_mode = 'static'
                flag_prefix = 'CTF{'
                flag_suffix = '}'
                random_flag_length = 0
                
                random_match = re.search(r'<ran_(\d+)>', flag_pattern)
                if random_match:
                    flag_mode = 'random'
                    random_flag_length = int(random_match.group(1))
                    parts = flag_pattern.split(random_match.group(0))
                    flag_prefix = parts[0] if len(parts) > 0 else 'CTF{'
                    flag_suffix = parts[1] if len(parts) > 1 else '}'
                else:
                    flag_prefix = flag_pattern
                    flag_suffix = ''
                
                # Parse scoring
                scoring_type = str(row_data.get('scoring_type', 'standard')).lower()
                value = None
                container_initial = None
                container_decay = None
                container_minimum = None
                decay_function = 'logarithmic'
                
                if scoring_type == 'dynamic':
                    container_initial = int(row_data.get('initial', 500))
                    container_decay = int(row_data.get('decay', 20))
                    container_minimum = int(row_data.get('minimum', 100))
                    decay_function = str(row_data.get('decay_function', 'logarithmic')).lower()
                    value = container_initial
                else:
                    value = int(row_data.get('value', 100))
                    container_initial = None
                    container_decay = None
                    container_minimum = None
                
                # Create challenge
                challenge = ContainerChallenge(
                    name=str(row_data['name']),
                    category=str(row_data['category']),
                    description=str(row_data.get('description', '')),
                    value=value,
                    state=str(row_data.get('state', 'visible')),
                    type='container',
                    
                    # Container fields
                    image=str(row_data['image']),
                    internal_port=int(row_data.get('internal_port', 22)),
                    command=str(row_data.get('command', '')),
                    container_connection_type=str(row_data.get('connection_type', 'ssh')),
                    container_connection_info=str(row_data.get('connection_info', '')),
                    
                    # Flag fields
                    flag_mode=flag_mode,
                    flag_prefix=flag_prefix,
                    flag_suffix=flag_suffix,
                    random_flag_length=random_flag_length,
                    
                    # Scoring fields
                    container_initial=container_initial,
                    container_decay=container_decay,
                    container_minimum=container_minimum,
                    decay_function=decay_function
                )
                
                db.session.add(challenge)
                db.session.flush()  # Get ID
                
                # Create dummy flag
                dummy_flag = Flags(
                    challenge_id=challenge.id,
                    type='static',
                    content='[Container flag - auto-generated per instance]',
                    data=''
                )
                db.session.add(dummy_flag)
                
                created_count += 1
                
            except Exception as e:
                errors.append(f'Row {row_idx}: {str(e)}')
                continue
        
        db.session.commit()
        
        # Return JSON response
        return jsonify({
            'success': True,
            'created': created_count,
            'errors': errors
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@admin_bp.route('/download-template')
@admins_only
def download_template():
    """Download CSV template for importing challenges"""
    try:
        from flask import make_response
        import io
        
        # CSV headers
        headers = [
            'name', 'category', 'description', 'image', 'internal_port',
            'command', 'connection_type', 'connection_info',
            'flag_pattern', 'scoring_type', 'value',
            'initial', 'decay', 'minimum', 'decay_function', 'state'
        ]
        
        # Example rows
        examples = [
            {
                'name': 'Web Challenge Example',
                'category': 'Web',
                'description': 'Find the flag',
                'image': 'nginx:latest',
                'internal_port': '80',
                'command': '',
                'connection_type': 'http',
                'connection_info': 'Access via browser',
                'flag_pattern': 'CTF{static_flag}',
                'scoring_type': 'standard',
                'value': '100',
                'initial': '',
                'decay': '',
                'minimum': '',
                'decay_function': '',
                'state': 'visible'
            },
            {
                'name': 'SSH Challenge Example',
                'category': 'Pwn',
                'description': 'SSH and find flag',
                'image': 'ubuntu:20.04',
                'internal_port': '22',
                'command': '/usr/sbin/sshd -D',
                'connection_type': 'ssh',
                'connection_info': 'user:ctf pass:ctf',
                'flag_pattern': 'CTF{<ran_16>}',
                'scoring_type': 'dynamic',
                'value': '',
                'initial': '500',
                'decay': '20',
                'minimum': '100',
                'decay_function': 'logarithmic',
                'state': 'visible'
            }
        ]
        
        # Build CSV
        output = io.StringIO()
        output.write(','.join(headers) + '\n')
        
        for example in examples:
            row = []
            for header in headers:
                value = example.get(header, '')
                # Escape commas and quotes
                if ',' in value or '"' in value:
                    value = '"' + value.replace('"', '""') + '"'
                row.append(value)
            output.write(','.join(row) + '\n')
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=container_challenges_template.csv'
        
        return response
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
